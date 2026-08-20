from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    Alignment,
    PatternFill,
    Border,
    Side
)
from openpyxl.utils import get_column_letter


class ExcelExporter:

    def __init__(self):

        self.colunas = [
            "Nome",
            "Número Processo",
            "Proc. Originário",
            "Vara",
            "Banco",
            "Nº RPV"
        ]

    # ============================================================
    # CONVERTER BANCO
    # ============================================================

    def formatar_banco(self, banco):

        if not banco:
            return ""

        banco = str(banco).strip().upper()

        if "BANCO DO BRASIL" in banco:
            return "BB"

        if "CAIXA ECONÔMICA FEDERAL" in banco:
            return "CEF"

        if "CAIXA ECONOMICA FEDERAL" in banco:
            return "CEF"

        if "CAIXA" in banco:
            return "CEF"

        if "BANCO DO BRASIL" in banco:
            return "BB"

        return banco

    # ============================================================
    # EXPORTAR
    # ============================================================

    def exportar(
        self,
        dados,
        caminho="relatorio_rpv.xlsx"
    ):

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "RPVs"

        # ========================================================
        # CORES
        # ========================================================

        cor_cabecalho = "1F4E78"
        cor_linha_par = "F3F6F9"
        cor_linha_impar = "FFFFFF"
        cor_borda = "D9E1F2"

        # ========================================================
        # ESTILOS
        # ========================================================

        fonte_cabecalho = Font(
            name="Calibri",
            size=11,
            bold=True,
            color="FFFFFF"
        )

        fonte_normal = Font(
            name="Calibri",
            size=11,
            color="000000"
        )

        fonte_link = Font(
            name="Calibri",
            size=11,
            color="0563C1",
            underline="single"
        )

        preenchimento_cabecalho = PatternFill(
            fill_type="solid",
            fgColor=cor_cabecalho
        )

        preenchimento_par = PatternFill(
            fill_type="solid",
            fgColor=cor_linha_par
        )

        preenchimento_impar = PatternFill(
            fill_type="solid",
            fgColor=cor_linha_impar
        )

        lado_borda = Side(
            style="thin",
            color=cor_borda
        )

        borda = Border(
            left=lado_borda,
            right=lado_borda,
            top=lado_borda,
            bottom=lado_borda
        )

        # ========================================================
        # CABEÇALHO
        # ========================================================

        for coluna, nome in enumerate(
            self.colunas,
            start=1
        ):

            celula = sheet.cell(
                row=1,
                column=coluna
            )

            celula.value = nome

            celula.font = fonte_cabecalho

            celula.fill = preenchimento_cabecalho

            celula.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            celula.border = borda

        # Altura do cabeçalho

        sheet.row_dimensions[1].height = 25

        # ========================================================
        # DADOS
        # ========================================================

        for linha, processo in enumerate(
            dados,
            start=2
        ):

            # ----------------------------------------------------
            # DADOS
            # ----------------------------------------------------

            nome = processo.get(
                "nome",
                ""
            )

            numero = processo.get(
                "numero",
                processo.get(
                    "processo",
                    ""
                )
            )

            processo_originario = processo.get(
                "processo_originario",
                ""
            )

            vara = processo.get(
                "vara",
                ""
            )

            banco = self.formatar_banco(
                processo.get(
                    "banco",
                    ""
                )
            )

            rpv = processo.get(
                "rpv",
                ""
            )

            link = processo.get(
                "link",
                ""
            )

            valores = [
                nome,
                numero,
                processo_originario,
                vara,
                banco,
                rpv
            ]

            # ----------------------------------------------------
            # PREENCHER LINHA
            # ----------------------------------------------------

            for coluna, valor in enumerate(
                valores,
                start=1
            ):

                celula = sheet.cell(
                    row=linha,
                    column=coluna
                )

                celula.value = valor

                celula.font = fonte_normal

                celula.border = borda

                celula.alignment = Alignment(
                    vertical="center"
                )

                # ------------------------------------------------
                # LINHAS ALTERNADAS
                # ------------------------------------------------

                if linha % 2 == 0:
                    celula.fill = preenchimento_par
                else:
                    celula.fill = preenchimento_impar

            # ----------------------------------------------------
            # PROCESSO COMO LINK
            # ----------------------------------------------------

            celula_processo = sheet.cell(
                row=linha,
                column=2
            )

            if numero and link:

                celula_processo.hyperlink = link

                celula_processo.font = fonte_link

                celula_processo.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            # ----------------------------------------------------
            # ALINHAMENTO CENTRAL
            # ----------------------------------------------------

            sheet.cell(
                row=linha,
                column=2
            ).alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            sheet.cell(
                row=linha,
                column=3
            ).alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            sheet.cell(
                row=linha,
                column=4
            ).alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            sheet.cell(
                row=linha,
                column=5
            ).alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            sheet.cell(
                row=linha,
                column=6
            ).alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            # Altura da linha

            sheet.row_dimensions[linha].height = 22

        # ========================================================
        # LARGURA DAS COLUNAS
        # ========================================================

        larguras = {
            "A": 35,  # Nome
            "B": 25,  # Número Processo
            "C": 25,  # Processo Originário
            "D": 18,  # Vara
            "E": 10,  # Banco
            "F": 20   # RPV
        }

        for coluna, largura in larguras.items():

            sheet.column_dimensions[
                coluna
            ].width = largura

        # ========================================================
        # CONGELAR CABEÇALHO
        # ========================================================

        sheet.freeze_panes = "A2"

        # ========================================================
        # FILTRO
        # ========================================================

        if sheet.max_row >= 1:

            sheet.auto_filter.ref = (
                f"A1:F{sheet.max_row}"
            )

        # ========================================================
        # CONFIGURAÇÕES DA PLANILHA
        # ========================================================

        sheet.sheet_view.showGridLines = False

        # ========================================================
        # TÍTULO / ABA
        # ========================================================

        sheet.sheet_properties.pageSetUpPr.fitToPage = True

        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0

        # ========================================================
        # SALVAR
        # ========================================================

        workbook.save(caminho)

        print(
            f"\n[OK] Excel salvo em: {caminho}"
        )

        # MUITO IMPORTANTE:
        # retornar o caminho para o ColetaView

        return caminho